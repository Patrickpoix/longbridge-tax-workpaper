from pathlib import Path
import json
from zipfile import ZipFile

from openpyxl import load_workbook

from longbridge_tax_workpaper.discovery import find_pdfs
from longbridge_tax_workpaper.runner import run_workpaper

from conftest import make_statement_pdf


def test_two_runs_do_not_reingest_output_pdfs(tmp_path: Path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    for month in range(1, 13):
        make_statement_pdf(input_dir / f"statement-monthly-2025{month:02d}-H00000001.pdf", f"2025{month:02d}")
    output_dir = input_dir / "outputs"
    first = run_workpaper(input_dir, output_dir, tax_year=2025, fx_rates={"USD": 7.0, "HKD": 0.9})
    first_hash = first.workbook.read_bytes()
    second = run_workpaper(input_dir, output_dir, tax_year=2025, fx_rates={"USD": 7.0, "HKD": 0.9})
    assert second.workbook.read_bytes() == first_hash
    assert len(find_pdfs(input_dir, exclude_roots=[output_dir])) == 12

    workbook = load_workbook(second.workbook, read_only=True, data_only=True)
    required = {
        "年度纳税汇总", "财产转让计税情景", "FIFO已实现盈亏", "移动平均已实现盈亏",
        "股息与预扣税", "融资利息应计", "融资利息实际支付", "持仓数量对账",
        "期初逐月持仓对账", "月度覆盖", "复核就绪性", "文件追溯", "版本信息",
    }
    assert required.issubset(set(workbook.sheetnames))
    with ZipFile(second.workpapers_zip) as archive:
        names = archive.namelist()
        assert not any("/source_pdfs/" in name for name in names)
        assert any(name.endswith("manifest.json") for name in names)

    with ZipFile(second.processed_delivery_zip) as archive:
        processed_readme = archive.read(
            next(name for name in archive.namelist() if name.endswith("/README.md"))
        ).decode("utf-8")
        assert "仍包含账户及交易级财务信息" in processed_readme

    with ZipFile(second.sanitized_delivery_zip) as archive:
        names = archive.namelist()
        assert len(names) == 4
        assert {Path(name).name for name in names} == {
            "README.md", "manifest.json", "review_status.json", "longbridge_2025_sanitized_review.xlsx",
        }
        status = json.loads(archive.read(next(name for name in names if name.endswith("review_status.json"))))
        assert set(status) == {"status", "review_status", "ready_to_file", "ready_for_review", "tax_year", "checks"}
        assert all("detail" not in item for item in status["checks"])
        manifest = json.loads(archive.read(next(name for name in names if name.endswith("manifest.json"))))
        assert manifest["package_version"] == "1.0.0"
        assert manifest["schema_version"] == "v4"
        xlsx_name = next(name for name in names if name.endswith(".xlsx"))
        extracted = tmp_path / "sanitized-review.xlsx"
        extracted.write_bytes(archive.read(xlsx_name))

    sanitized_workbook = load_workbook(extracted, read_only=True, data_only=True)
    assert set(sanitized_workbook.sheetnames) == {
        "年度纳税汇总", "财产转让计税情景", "年末汇率", "复核就绪性", "版本信息",
    }
    summary = sanitized_workbook["年度纳税汇总"]
    account_values = [row[1].value for row in summary.iter_rows(min_row=1, max_row=12) if row[0].value == "账户"]
    assert account_values == ["已脱敏"]
    sanitized_text = "\n".join(
        str(cell.value)
        for sheet in sanitized_workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "H00000001" not in sanitized_text
    for source_pdf in input_dir.glob("*.pdf"):
        assert source_pdf.name not in sanitized_text

    archival = run_workpaper(
        input_dir,
        tmp_path / "archival-output",
        tax_year=2025,
        fx_rates={"USD": 7.0, "HKD": 0.9},
        include_source_pdfs=True,
    )
    with ZipFile(archival.workpapers_zip) as archive:
        pdf_names = [name for name in archive.namelist() if "/source_pdfs/" in name and name.endswith(".pdf")]
        assert len(pdf_names) == 12
