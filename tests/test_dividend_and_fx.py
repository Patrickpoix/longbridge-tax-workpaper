from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from longbridge_tax_workpaper.config import prepare_runtime_config, runtime_config_environment
from longbridge_tax_workpaper.cost_basis import MethodResult
from longbridge_tax_workpaper.dividends import build_dividend_tax_basis_rows
from longbridge_tax_workpaper.filing_readiness import assess_filing_readiness
from longbridge_tax_workpaper.reporting import build_processed_workbook
from longbridge_tax_workpaper.schema import FieldValue, SectionResult, StatementResult


def fv(value):
    return FieldValue.native(value, confidence=1.0)


def dividend_statement(detail: str, *, amount: float = 18.0, currency: str = "HKD") -> StatementResult:
    st = StatementResult("202501", "sample.pdf")
    st.sections["other_fund_flows"] = SectionResult("other_fund_flows", rows=[{
        "tax_category": fv("dividend_income"),
        "raw_detail": fv(detail),
        "currency": fv(currency),
        "cash_amount": fv(amount),
        "date": fv("2025.01.15"),
    }])
    return st


@pytest.mark.parametrize("detail", [
    "#1288.HK DIVIDEND RMB0.20/SH Held:100 (-10%)",
    "#1288.HK 股息 人民币：0.20 每股，持股：100（－10％）",
    "#1288.HK DIVIDEND CNY 0.20 / 股 持有 100 - 10%",
])
def test_embedded_withholding_regex_variants(tmp_path: Path, detail: str):
    paths = prepare_runtime_config(tmp_path / "config", tax_year=2025, account_opening_month="202501", fx_rates={"USD": 7, "HKD": 0.9})
    with runtime_config_environment(paths):
        row = build_dividend_tax_basis_rows([dividend_statement(detail)])[0]
    assert row["gross_dividend_cny"] == 20.0
    assert row["embedded_withholding_cny"] == 2.0
    assert row["automatic_credit_cny"] == 0.0
    assert row["statement_withholding_credit_candidate_cny"] == 2.0


def test_direct_us_withholding_and_no_double_consumption(tmp_path: Path):
    st = StatementResult("202501", "sample.pdf")
    st.sections["other_fund_flows"] = SectionResult("other_fund_flows", rows=[
        {"tax_category": fv("dividend_income"), "raw_detail": fv("#AVGO.US Cash Dividend"), "currency": fv("USD"), "cash_amount": fv(9), "date": fv("2025.01.15")},
        {"tax_category": fv("withholding_tax"), "raw_detail": fv("#AVGO.US Withholding Tax"), "currency": fv("USD"), "cash_amount": fv(-1), "date": fv("2025.01.15")},
        {"tax_category": fv("dividend_income"), "raw_detail": fv("#AVGO.US Cash Dividend second line"), "currency": fv("USD"), "cash_amount": fv(4), "date": fv("2025.01.15")},
    ])
    paths = prepare_runtime_config(tmp_path / "config", tax_year=2025, account_opening_month="202501", fx_rates={"USD": 7, "HKD": 0.9})
    with runtime_config_environment(paths):
        rows = build_dividend_tax_basis_rows([st])
    assert rows[0]["filing_dividend_income_cny"] == 63.0
    assert rows[0]["direct_withholding_cny"] == 7.0
    assert rows[0]["statement_withholding_credit_candidate_cny"] == 7.0
    assert rows[1]["direct_withholding_cny"] == 0.0
    assert rows[0]["automatic_credit_cny"] == 0.0


def test_missing_fx_marks_dividend_and_workbook_incomplete(tmp_path: Path):
    statements = [StatementResult(f"2025{m:02d}", f"{m}.pdf") for m in range(1, 13)]
    statements[0].sections["other_fund_flows"] = dividend_statement("#AVGO.US Cash Dividend", amount=9, currency="USD").sections["other_fund_flows"]
    report = {"fifo": MethodResult(method="FIFO"), "moving_average": MethodResult(method="MOVING_AVERAGE"), "differences": [], "opening_lots": [], "prior_period_coverage": {"status": "ok"}, "errors": []}
    paths = prepare_runtime_config(tmp_path / "config", tax_year=2025, account_opening_month="202501", fx_rates={})
    with runtime_config_environment(paths):
        dividends = build_dividend_tax_basis_rows(statements)
        readiness = assess_filing_readiness(statements, cost_report=report)
        target = build_processed_workbook(
            tmp_path / "result.xlsx", tax_year=2025, account_id="H1", statements=statements,
            prior_statements=[], cost_report=report, dividends=dividends, margin_accrual=[], margin_actual=[],
            readiness=readiness, source_files=[],
        )
    assert dividends[0]["filing_dividend_income_cny"] is None
    assert dividends[0]["cny_conversion_status"] == "incomplete_missing_fx"
    fx_check = next(row for row in readiness["checks"] if row["code"] == "YEAR_END_FX")
    assert fx_check["status"] == "BLOCKED"
    assert readiness["status"] == "BLOCKED_FOR_REVIEW"
    wb = load_workbook(target, data_only=True)
    summary_values = [cell.value for row in wb["年度纳税汇总"].iter_rows() for cell in row]
    assert "缺少年末汇率" in summary_values
    assert 0 not in [dividends[0]["filing_dividend_income_cny"]]
