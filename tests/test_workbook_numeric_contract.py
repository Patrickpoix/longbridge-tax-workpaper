from pathlib import Path

from openpyxl import load_workbook

from longbridge_tax_workpaper.config import prepare_runtime_config, runtime_config_environment
from longbridge_tax_workpaper.cost_basis import MethodResult
from longbridge_tax_workpaper.reporting import build_processed_workbook
from longbridge_tax_workpaper.schema import StatementResult


def disposal(ref: str, market: str, pnl: float):
    return {
        "source_reference": ref,
        "trade_date": "2025-01-01",
        "security_id": f"{market}:{ref}",
        "symbol": ref,
        "asset_category": "stock",
        "market": market,
        "currency": "HKD" if market == "HK" else "USD",
        "quantity": 1,
        "gross_proceeds": 100,
        "disposal_fees": 1,
        "net_proceeds": 99,
        "allocated_cost": 99 - pnl,
        "realized_pnl": pnl,
        "realized_pnl_cny": pnl,
        "cny_conversion_status": "complete",
        "validation_status": "ok",
    }


def sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows) if row and row[0] == "所得类别")
    headers = rows[header_index]
    return [dict(zip(headers, row)) for row in rows[header_index + 1:] if row and row[0]]


def test_workbook_summary_keeps_numeric_zero_and_reconciles_scenarios(tmp_path: Path):
    statements = [StatementResult(f"2025{month:02d}", f"{month}.pdf") for month in range(1, 13)]
    fifo = MethodResult(method="FIFO", disposals=[disposal("A", "HK", 100), disposal("B", "US", -40)])
    moving = MethodResult(method="MOVING_AVERAGE", disposals=[disposal("A", "HK", 80), disposal("B", "US", -20)])
    report = {"fifo": fifo, "moving_average": moving, "differences": [], "opening_lots": []}
    dividends = [{
        "statement_month": "202501",
        "filing_dividend_income_cny": 100.0,
        "statement_withholding_credit_candidate_cny": 10.0,
        "automatic_credit_cny": 0.0,
        "cny_conversion_status": "complete",
    }]
    paths = prepare_runtime_config(tmp_path / "config", tax_year=2025, account_opening_month="202501", fx_rates={"USD": 7, "HKD": 0.9})
    target = tmp_path / "workbook.xlsx"
    with runtime_config_environment(paths):
        build_processed_workbook(
            target, tax_year=2025, account_id="H1", statements=statements, prior_statements=[],
            cost_report=report, dividends=dividends, margin_accrual=[], margin_actual=[],
            readiness={"status": "REVIEW_REQUIRED", "ready_to_file": False, "checks": []}, source_files=[],
        )
    wb = load_workbook(target, data_only=True)
    records = sheet_records(wb["年度纳税汇总"])
    dividend = next(row for row in records if row["所得类别"] == "股息红利所得")
    assert dividend["应纳税所得额(CNY)"] == 100
    assert dividend["参考税额(CNY)"] == 20
    assert dividend["自动抵免(CNY)"] == 0
    assert dividend["抵免候选(CNY)"] == 10

    fifo_cross = next(row for row in records if row["计算口径"] == "先进先出法（FIFO） - 同一账户跨市场合并净额")
    fifo_separate = next(row for row in records if row["计算口径"] == "先进先出法（FIFO） - 各市场分别计算后合计")
    fifo_positive = next(row for row in records if row["计算口径"] == "先进先出法（FIFO） - 逐笔正收益、不抵减亏损")
    assert (fifo_cross["应纳税所得额(CNY)"], fifo_cross["参考税额(CNY)"]) == (60, 12)
    assert (fifo_separate["应纳税所得额(CNY)"], fifo_separate["参考税额(CNY)"]) == (100, 20)
    assert (fifo_positive["应纳税所得额(CNY)"], fifo_positive["参考税额(CNY)"]) == (100, 20)
