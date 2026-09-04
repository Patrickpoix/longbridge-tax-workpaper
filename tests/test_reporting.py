from pathlib import Path

from openpyxl import load_workbook

from longbridge_tax_workpaper.cost_basis import MethodResult
from longbridge_tax_workpaper.reporting import (
    SANITIZED_SHEETS,
    build_processed_workbook,
    build_sanitized_review_workbook,
)
from longbridge_tax_workpaper.schema import StatementResult


def test_single_workbook_multiple_sheets(tmp_path: Path):
    statement = StatementResult(statement_month="202701", source_pdf="sample.pdf")
    report = {"fifo": MethodResult(method="FIFO"), "moving_average": MethodResult(method="MOVING_AVERAGE"), "differences": [], "opening_lots": []}
    target = tmp_path / "result.xlsx"
    build_processed_workbook(
        target, tax_year=2027, account_id="H123", statements=[statement], prior_statements=[], cost_report=report,
        dividends=[], margin_accrual=[], margin_actual=[], readiness={"status": "READY_FOR_REVIEW", "ready_to_file": False, "checks": []}, source_files=[],
    )
    workbook = load_workbook(target, read_only=True, data_only=False)
    assert "年度纳税汇总" in workbook.sheetnames
    assert "FIFO已实现盈亏" in workbook.sheetnames
    assert "移动平均已实现盈亏" in workbook.sheetnames
    assert "复核就绪性" in workbook.sheetnames
    assert len(workbook.sheetnames) > 10


def test_sanitized_workbook_keeps_only_aggregate_sheets_and_redacts_detail(tmp_path: Path):
    statement = StatementResult(statement_month="202701", source_pdf="private-source.pdf")
    report = {"fifo": MethodResult(method="FIFO"), "moving_average": MethodResult(method="MOVING_AVERAGE"), "differences": [], "opening_lots": []}
    full = tmp_path / "full.xlsx"
    sanitized = tmp_path / "sanitized.xlsx"
    build_processed_workbook(
        full, tax_year=2027, account_id="H00000001", statements=[statement], prior_statements=[], cost_report=report,
        dividends=[], margin_accrual=[], margin_actual=[],
        readiness={
            "status": "REVIEW_REQUIRED",
            "ready_to_file": False,
            "checks": [{"code": "X", "label": "检查", "status": "WARNING", "blocking": False, "risk_type": "technical", "detail": "private-source.pdf"}],
        },
        source_files=[{"文件名": "private-source.pdf", "SHA-256": "f" * 64}],
    )
    build_sanitized_review_workbook(full, sanitized)
    workbook = load_workbook(sanitized, read_only=True, data_only=True)
    assert set(workbook.sheetnames) == SANITIZED_SHEETS
    summary = workbook["年度纳税汇总"]
    account_values = [row[1].value for row in summary.iter_rows(min_row=1, max_row=12) if row[0].value == "账户"]
    assert account_values == ["已脱敏"]
    readiness = workbook["复核就绪性"]
    values = [cell.value for row in readiness.iter_rows() for cell in row]
    assert "private-source.pdf" not in values
    assert "详见完整底稿" in values
