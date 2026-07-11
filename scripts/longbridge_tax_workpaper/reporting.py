from __future__ import annotations

import json
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .filing_policy import load_tax_policy, year_end_fx_rate
from .money import decimal_value, q_cny, to_float
from .schema import StatementResult
from .serialization import section_rows
from .xlsx_determinism import canonicalize_xlsx_package

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FONT = Font(bold=True, color="17365D", size=15)
SUBTITLE_FILL = PatternFill("solid", fgColor="EAF2F8")
SUBTITLE_FONT = Font(bold=True, color="17365D")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")

ZH_HEADERS = {
    "statement_month": "月结单月份", "source_pdf": "来源PDF", "row_index": "行号",
    "method": "成本方法", "trade_date": "交易日期", "execution_time": "成交时间",
    "source_reference": "成交/事件编号", "event_type": "事件类型", "security_id": "证券标识",
    "symbol": "证券名称/合约", "asset_category": "资产类别", "market": "交易市场",
    "currency": "币种", "quantity": "数量", "gross_proceeds": "处置收入总额",
    "disposal_fees": "卖出/处置费用", "net_proceeds": "处置净收入",
    "allocated_cost": "分配成本", "realized_pnl": "已实现盈亏",
    "year_end_fx_rate": "年末人民币中间价", "year_end_cny_rate": "年末人民币中间价",
    "realized_pnl_cny": "已实现盈亏(CNY)", "cny_conversion_status": "人民币折算状态",
    "validation_status": "校验状态", "validation_note": "校验说明", "match_detail": "成本匹配明细",
    "match_detail_json": "成本匹配明细", "source_type": "来源类型",
    "source_pdf_sha256": "来源PDF SHA-256", "total_cost": "总成本",
    "unit_cost": "单位成本", "acquired_date": "取得日期", "acquired_time": "取得时间",
    "evidence": "证据", "fifo_allocated_cost": "FIFO分配成本",
    "moving_average_allocated_cost": "移动加权平均分配成本",
    "fifo_realized_pnl": "FIFO已实现盈亏", "moving_average_realized_pnl": "移动加权平均已实现盈亏",
    "pnl_difference": "盈亏差异", "fifo_realized_pnl_cny": "FIFO已实现盈亏(CNY)",
    "moving_average_realized_pnl_cny": "移动加权平均已实现盈亏(CNY)",
    "pnl_difference_cny": "盈亏差异(CNY)", "computed_ending_quantity": "成本账期末数量",
    "statement_ending_quantity": "月结单期末数量", "difference": "差异",
    "filing_dividend_income_cny": "股息申报收入(CNY)",
    "china_tax_before_credit_cny": "中国税额参考(CNY)",
    "statement_withholding_credit_candidate_cny": "月结单扣税抵免候选(CNY)",
    "automatic_credit_cny": "自动抵免(CNY)", "gross_dividend_amount": "税前股息",
    "cash_received_amount": "现金到账", "cash_received": "现金到账",
    "withholding_tax_amount": "预扣税", "withholding_rate": "预扣税率",
    "withholding_mode": "扣税方式", "security_code": "证券代码",
    "security_name": "证券名称", "month_label": "月份",
    "hkd_accrued_interest": "港元应计融资利息", "usd_accrued_interest": "美元应计融资利息",
    "pdf_usd_hkd_reference_rate": "PDF月内USD/HKD汇率",
    "usd_interest_hkd_equivalent": "美元利息折港元",
    "total_margin_interest_hkd_tax_basis": "融资利息合计(HKD)",
    "total_actual_payment_hkd_equivalent": "实际支付折港元合计", "raw_detail": "月结单原文",
    "tax_category": "税务分类", "cash_amount": "现金金额", "amount": "金额", "date": "日期",
    "name": "名称", "opening_position": "期初数量", "ending_position": "期末数量",
    "cost": "券商展示成本", "price": "价格", "total_amount": "结算总额", "order_id": "订单号",
    "direction": "方向", "side": "买卖方向", "status": "状态", "blocking": "是否阻断",
    "detail": "说明", "code": "检查代码", "label": "检查项目",
    "source_status": "来源状态", "source_date": "来源日期", "source_url": "来源网址",
    "evidence_sha256": "证据SHA-256", "rate": "汇率", "unit": "单位",
}

VALUE_ZH = {
    "PASS": "通过", "FAIL": "失败", "WARNING": "警告", "BLOCKED": "阻断",
    "TECHNICALLY_GENERATED": "技术生成完成", "REVIEW_REQUIRED": "需要复核",
    "BLOCKED_FOR_REVIEW": "复核阻断", "READY_FOR_REVIEW": "可开始复核",
    "REVIEW_BLOCKED": "复核阻断", "ok": "通过", "error": "错误",
    "complete": "完整", "incomplete_missing_fx": "缺少年末汇率",
    "stock": "股票", "option": "期权", "warrant": "权证",
    "BUY": "买入", "SELL": "卖出", "TRADE_SELL": "普通卖出", "AUTO_EX": "到期自动结算",
    "HK": "香港市场", "US": "美国市场", "UNKNOWN": "未知",
    "FIFO": "先进先出法（FIFO）", "MOVING_AVERAGE": "移动加权平均法",
    True: "是", False: "否",
}

_TEXT_HEAVY_HEADERS = {
    "evidence", "raw_detail", "match_detail", "match_detail_json", "validation_note",
    "detail", "source_url", "cash_flow_evidence", "basis_note", "note",
}


def _plain(value: Any) -> Any:
    # bool is a subclass of int in Python; test it before string label lookup so
    # numeric 0/1 are not rendered as 否/是 in financial cells.
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, str) and value in VALUE_ZH:
        return VALUE_ZH[value]
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _headers(rows: list[dict[str, Any]], preferred: list[str] | None = None) -> list[str]:
    result: list[str] = []
    for key in preferred or []:
        if any(key in row for row in rows):
            result.append(key)
    for row in rows:
        for key in row:
            if key not in result:
                result.append(key)
    return result


def _style_header(cells: Iterable[Any]) -> None:
    for cell in cells:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_table(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    *,
    title: str | None = None,
    note: str | None = None,
    preferred_headers: list[str] | None = None,
) -> None:
    sheet = wb.create_sheet(sheet_name[:31])
    headers = _headers(rows, preferred_headers)
    if not headers:
        headers = ["说明"]
        rows = [{"说明": "无数据"}]
    row_num = 1
    if title:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = sheet.cell(1, 1, title)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = Alignment(vertical="center")
        row_num += 1
    if note:
        sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        cell = sheet.cell(row_num, 1, note)
        cell.fill = NOTE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_num += 1
    header_row = row_num
    for column, header in enumerate(headers, start=1):
        sheet.cell(header_row, column, ZH_HEADERS.get(header, header))
    _style_header(sheet[header_row])
    for record in rows:
        sheet.append([_plain(record.get(header)) for header in headers])
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{sheet.max_row}"
    for column, header in enumerate(headers, start=1):
        sample = [str(record.get(header) or "") for record in rows[:200]]
        max_width = 80 if header in _TEXT_HEAVY_HEADERS else 40
        width = min(max_width, max([10, len(ZH_HEADERS.get(header, header)) + 2, *[len(value) + 2 for value in sample]]))
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _all_complete(values: list[object]) -> bool:
    return all(value not in (None, "") for value in values)


def _sum_complete(rows: list[dict[str, Any]], key: str) -> Decimal | None:
    values = [decimal_value(row.get(key)) for row in rows]
    if any(value is None for value in values):
        return None
    return q_cny(sum((value for value in values if value is not None), Decimal("0")))


def _summary_scenarios(cost_report: dict[str, Any]) -> list[dict[str, Any]]:
    scenarios: list[dict[str, Any]] = []
    for method_name, result_key in (("FIFO", "fifo"), ("MOVING_AVERAGE", "moving_average")):
        result = cost_report.get(result_key)
        disposals = list(getattr(result, "disposals", []))
        cny_values = [decimal_value(row.get("realized_pnl_cny")) for row in disposals]
        complete = all(value is not None for value in cny_values)
        valid = [value for value in cny_values if value is not None]
        if complete:
            total = sum(valid, Decimal("0"))
            positive = sum((max(value, Decimal("0")) for value in valid), Decimal("0"))
            by_market: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
            for row, value in zip(disposals, valid):
                by_market[str(row.get("market") or "UNKNOWN")] += value
            separate_taxable = sum((max(value, Decimal("0")) for value in by_market.values()), Decimal("0"))
            scenario_values: list[tuple[str, Decimal | None, str]] = [
                ("同一账户跨市场合并净额", max(total, Decimal("0")), "同一长桥账户内各市场全年已实现盈亏合并，仅作测算情景。"),
                ("各市场分别计算后合计", separate_taxable, "各市场内部净额为正的部分相加，仅作测算情景。"),
                ("逐笔正收益、不抵减亏损", positive, "只汇总正收益处置，亏损不抵减，仅作保守对照。"),
            ]
        else:
            scenario_values = [
                ("同一账户跨市场合并净额", None, "缺少年末汇率，人民币情景未计算。"),
                ("各市场分别计算后合计", None, "缺少年末汇率，人民币情景未计算。"),
                ("逐笔正收益、不抵减亏损", None, "缺少年末汇率，人民币情景未计算。"),
            ]
        for label, taxable, note in scenario_values:
            taxable_q = q_cny(taxable) if taxable is not None else None
            tax = q_cny(taxable_q * Decimal("0.20")) if taxable_q is not None else None
            scenarios.append({
                "成本方法": VALUE_ZH[method_name],
                "情景": label,
                "应纳税所得额(CNY)": to_float(taxable_q),
                "参考税率": 0.20,
                "参考税额(CNY)": to_float(tax),
                "人民币折算状态": "完整" if taxable_q is not None else "缺少年末汇率",
                "说明": note,
            })
    return scenarios


def _reward_cny(statements: list[StatementResult], policy: dict[str, Any]) -> tuple[Decimal | None, str]:
    values: list[Decimal] = []
    for row in section_rows(statements, "other_fund_flows"):
        if row.get("tax_category") != "cash_reward_other_income":
            continue
        currency = str(row.get("currency") or "")
        rate = year_end_fx_rate(currency, policy) if currency in {"HKD", "USD"} else None
        if rate is None:
            return None, "缺少年末汇率"
        amount = decimal_value(row.get("cash_amount") if row.get("cash_amount") is not None else row.get("amount"), default=Decimal("0"))
        values.append(amount * decimal_value(rate))
    return q_cny(sum(values, Decimal("0"))), "完整"


def _fx_rows(policy: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for currency in ("USD", "HKD"):
        item = dict(policy.get("year_end_fx_rates", {}).get(currency, {}))
        rows.append({"币种": currency, **item})
    return rows


def build_processed_workbook(
    path: str | Path,
    *,
    tax_year: int,
    account_id: str | None,
    statements: list[StatementResult],
    prior_statements: list[StatementResult],
    cost_report: dict[str, Any],
    dividends: list[dict[str, Any]],
    margin_accrual: list[dict[str, Any]],
    margin_actual: list[dict[str, Any]],
    readiness: dict[str, Any],
    source_files: list[dict[str, Any]],
) -> Path:
    policy = load_tax_policy()
    wb = Workbook()
    summary = wb.active
    summary.title = "年度纳税汇总"
    summary.merge_cells("A1:H1")
    summary["A1"] = f"长桥证券 {tax_year} 年度税务工作底稿"
    summary["A1"].fill = TITLE_FILL
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A2:H2")
    summary["A2"] = "本文件为中国内地税收居民、单一长桥账户的税务工作底稿，不等同于税务机关认可的最终申报表。"
    summary["A2"].fill = NOTE_FILL
    summary["A2"].alignment = Alignment(wrap_text=True)
    meta = [
        ("纳税年度", tax_year), ("账户", account_id or "未识别"),
        ("月结单月份数", len(statements)), ("工作底稿状态", readiness.get("status")),
    ]
    for idx, (label, value) in enumerate(meta, start=3):
        summary.cell(idx, 1, label).fill = SUBTITLE_FILL
        summary.cell(idx, 1).font = SUBTITLE_FONT
        summary.cell(idx, 2, _plain(value))

    dividend_income = _sum_complete(dividends, "filing_dividend_income_cny")
    dividend_candidate = _sum_complete(dividends, "statement_withholding_credit_candidate_cny")
    reward_cny, reward_status = _reward_cny(statements, policy)
    scenario_rows = _summary_scenarios(cost_report)

    start = 8
    headers = ["所得类别", "计算口径", "应纳税所得额(CNY)", "参考税率", "参考税额(CNY)", "自动抵免(CNY)", "抵免候选(CNY)", "状态"]
    for col, header in enumerate(headers, start=1):
        summary.cell(start, col, header)
    _style_header(summary[start])

    dividend_tax = q_cny(dividend_income * Decimal("0.20")) if dividend_income is not None else None
    reward_tax = q_cny(reward_cny * Decimal("0.20")) if reward_cny is not None else None
    rows: list[list[Any]] = [
        ["股息红利所得", "月结单税前股息年度折算", to_float(dividend_income), 0.20, to_float(dividend_tax), 0.0, to_float(dividend_candidate), "工作底稿" if dividend_income is not None else "缺少年末汇率"],
        ["现金奖励", "其他/偶然性质收入候选", to_float(reward_cny), 0.20, to_float(reward_tax), 0.0, 0.0, "工作底稿" if reward_status == "完整" else reward_status],
    ]
    for item in scenario_rows:
        rows.append([
            "财产转让所得", f"{item['成本方法']} - {item['情景']}", item["应纳税所得额(CNY)"],
            0.20, item["参考税额(CNY)"], 0.0, 0.0,
            "测算情景" if item["人民币折算状态"] == "完整" else item["人民币折算状态"],
        ])
    for row_data in rows:
        summary.append(row_data)
    for row_num in range(start + 1, start + 1 + len(rows)):
        summary.cell(row_num, 4).number_format = "0.00%"
        for col in range(3, 8):
            summary.cell(row_num, col).number_format = "#,##0.00"
    summary.freeze_panes = f"A{start + 1}"
    summary.auto_filter.ref = f"A{start}:H{start + len(rows)}"
    for idx, width in enumerate([18, 40, 20, 12, 18, 18, 18, 16], 1):
        summary.column_dimensions[get_column_letter(idx)].width = width

    _write_table(wb, "财产转让计税情景", scenario_rows, title="财产转让所得测算情景", note="不同情景并列展示，不自动认定唯一申报口径。缺少年末汇率时人民币金额留空。")
    _write_table(wb, "FIFO已实现盈亏", list(getattr(cost_report.get("fifo"), "disposals", [])), title="先进先出法（FIFO）逐笔已实现盈亏")
    _write_table(wb, "移动平均已实现盈亏", list(getattr(cost_report.get("moving_average"), "disposals", [])), title="移动加权平均法逐笔已实现盈亏")
    _write_table(wb, "盈亏方法差异", list(cost_report.get("differences", [])), title="两种成本方法逐笔差异")
    remaining: list[dict[str, Any]] = []
    reconciliation: list[dict[str, Any]] = []
    for key in ("fifo", "moving_average"):
        result = cost_report.get(key)
        remaining.extend(list(getattr(result, "remaining_lots", [])))
        reconciliation.extend(list(getattr(result, "reconciliation", [])))
    _write_table(wb, "年末剩余批次", remaining, title="年末剩余成本批次", note="未实现盈亏不纳入年度已实现盈亏。")
    _write_table(wb, "持仓数量对账", reconciliation, title="成本账与月结单期末持仓数量对账")
    _write_table(
        wb,
        "期初逐月持仓对账",
        list(cost_report.get("prior_period_coverage", {}).get("monthly_reconciliation", [])),
        title="税年前历史月份逐月持仓滚动对账",
        note="逐月重放成交与公司行动，并与对应月份月结单期末持仓核对。",
    )
    _write_table(wb, "期初成本批次", list(cost_report.get("opening_lots", [])), title="纳税年度期初成本批次")
    _write_table(wb, "股息与预扣税", dividends, title="股息、现金到账与预扣税明细", note="自动抵免默认0；月结单扣税单列为抵免候选。")
    _write_table(wb, "融资利息应计", margin_accrual, title="融资利息月末应计口径", note="USD融资利息按每月PDF列示汇率折为HKD。")
    _write_table(wb, "融资利息实际支付", margin_actual, title="融资利息资金流水实际支付口径")
    _write_table(wb, "全部交易", section_rows(statements, "stock_trades") + section_rows(statements, "option_trades"), title="纳税年度全部证券交易")
    _write_table(wb, "全部资金流水", section_rows(statements, "other_fund_flows"), title="纳税年度全部资金流水")
    _write_table(wb, "全部持仓", section_rows(statements, "holdings"), title="各月持仓记录")
    monthly_rows = [{
        "月份": statement.statement_month, "来源PDF": Path(statement.source_pdf).name,
        "股票交易数": len(statement.sections.get("stock_trades").rows if statement.sections.get("stock_trades") else []),
        "期权交易数": len(statement.sections.get("option_trades").rows if statement.sections.get("option_trades") else []),
        "资金流水数": len(statement.sections.get("other_fund_flows").rows if statement.sections.get("other_fund_flows") else []),
        "阻断级错误数": sum(1 for value in statement.validations if value.severity == "error" and not value.passed),
    } for statement in statements]
    _write_table(wb, "月度覆盖", monthly_rows, title=f"{tax_year}年月结单覆盖")
    _write_table(wb, "年末汇率", _fx_rows(policy), title=f"{tax_year}-12-31人民币汇率与证据", note="缺失汇率不会被替换为0，相关人民币输出留空并在复核状态中阻断。")
    _write_table(wb, "复核就绪性", list(readiness.get("checks", [])), title="工作底稿复核就绪性检查")
    _write_table(wb, "文件追溯", source_files, title="输入文件与SHA-256追溯")
    _write_table(wb, "版本信息", [
        {"组件": "解析器", "版本": "longbridge-tax-workpaper-parser-v4"},
        {"组件": "数据结构", "版本": "longbridge-tax-workpaper-schema-v4"},
        {"组件": "工作簿", "版本": "openpyxl-workbook-v4"},
        {"组件": "精度规则", "版本": "Decimal中间8位；CNY输出2位；ROUND_HALF_UP"},
        {"组件": "适用范围", "版本": "中国内地税收居民/单一长桥账户/税务工作底稿"},
    ], title="版本与适用范围")

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    canonicalize_xlsx_package(target)
    return target
